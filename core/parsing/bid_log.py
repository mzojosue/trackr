from datetime import datetime
import hashlib
import os
import unicodedata
from parse import parse

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, colors

from core import environment
from core.log import logger

today = datetime.today


# Row Color Styling
submitted_bid = PatternFill(fill_type='solid', start_color=colors.YELLOW, end_color=colors.YELLOW)
cancelled_bid = PatternFill(fill_type='solid', start_color=colors.RED, end_color=colors.RED)

# Font Styling
base_font = Font(name="Calibri Light", size=12, color=colors.BLACK)
bold_font = base_font.copy(bold=True)

borders = Border(left   = Side(border_style=None, color='BBFFFFFF'),  # light gray border
				 right  = Side(border_style=None, color='BBFFFFFF'),
				 top    = Side(border_style=None, color='BBFFFFFF'),
				 bottom = Side(border_style=None, color='BBFFFFFF'))


def check_estimating_log(estimating_log=environment.get_estimating_log):
	""" Checks to see that content in database is the updated based on the stored hash of po_log file """
	md5 = hashlib.md5()
	with open(estimating_log, 'rb') as _estimating_log:
		buf = _estimating_log.read()
		md5.update(buf)
		_hash = str(md5.hexdigest())
		if _hash != environment.last_estimating_log_hash:
			logger.info('updating Estimating Log hash digest stored')
			environment.set_estimating_log_hash(_hash)
			return False
		return True


def parse_est_log(estimating_log=environment.get_estimating_log):
	log = load_workbook(estimating_log, read_only=True)
	_sheet = log.get_active_sheet()
	logger.debug('Opening Estimating Log')

	_prev = None  # buffer for storing previous bids. Used for grouping and alternate bidders
	for _row in _sheet.rows:
		__num = _row[0].value
		print "Processing bid row %s" % __num

		try:
			__num = int(__num)
			__name = unicodedata.normalize('NFKD', _row[1].value).encode('ASCII', 'ignore')
		except (ValueError, TypeError):
			# attempt to parse _row as a sub bid
			if _row[5].value and __num is None:  # sub bid hash is based off of 'gc' string/column and 'number' should be None
				_attr = (None, None, 'date_received', 'bid_date', None, 'gc', 'gc_contact' , None, 'scope')
				sub_bid = {}  # stores grabbed values
				for attr in _attr:
					if attr:
						_col = _attr.index(attr)
						sub_bid[attr] = _row[_col].value

				for i in ('date_received', 'bid_date'):
					__date = sub_bid[i]
					if "@" in str(__date):
						__date_due = [i for i in parse("{} @ {}", __date)]
					# create datetime object
					else:
						__date = [__date]
					_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
					for _format in _date_formats:
						try:
							__due = datetime.strptime(__date[0], _format)
							if __date[1:]:
								#TODO: implement
								#TODO: translate AM/PM
								#__due.hour = parse("{:d}{}", __date_due[1])[0]
								pass
							sub_bid[i] = __due
							break
						except ValueError:
							continue
						except TypeError:
							break
					yield 'sub_bid', (sub_bid, _prev['job_num'])
			continue  # move onto next row

		__date_recvd = _row[2].value
		if __date_recvd is not None:
			_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
			for _format in _date_formats:
				try:
					__date_recvd = datetime.strptime(__date_recvd, _format)
					break
				except ValueError:
					continue
				except TypeError:
					break

		# Parse Bid Due Date #
		__date_due = _row[3].value
		if __date_due is None:
			__date_due = 'ASAP'
		elif __date_due == 'ASAP':
			__date_due = str('ASAP')
		elif __date_due != 'ASAP':
			try:
				__date_due = _row[3].value
				__date_due.date()
			except AttributeError:
				if "@" in str(__date_due):
					__date_due = [i for i in parse("{} @ {}", __date_due)]
				# create datetime object
				else:
					__date_due = [__date_due]
				_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
				for _format in _date_formats:
					try:
						__due = datetime.strptime(__date_due[0], _format)
						if __date_due[1:]:
							#TODO: implement
							#TODO: translate AM/PM
							#__due.hour = parse("{:d}{}", __date_due[1])[0]
							pass
						__date_due = __due
						break
					except ValueError:
						continue
					except TypeError:
						break
		# Parse Date Submitted #
		__date_sent = _row[4].value
		if hasattr(__date_sent, 'lower') and __date_sent.lower() == "no bid":
			__date_sent = "No bid"
		elif __date_sent is not None:
			_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
			for _format in _date_formats:
				try:
					__date_sent = datetime.strptime(__date_sent, _format)
					break
				except ValueError:
					continue
				except TypeError:
					break

		if _row[5].value: __gc = _row[5].value          # Default: None
		else: __gc = None
		if _row[6].value: __gc_contact = _row[6].value  # Default: None
		else: __gc_contact = None
		#  __via = _row[7].value# default: email

		# Scope parsing #
		__scope = str(_row[8].value)
		if ',' in __scope or len(__scope) == 1:
			_tmp_scope = []
			valid_scope = ('M', 'E', 'I', 'B', 'P')
			for _letter in __scope:
				if _letter in valid_scope:
					_tmp_scope.append(_letter)
			__scope = _tmp_scope   # does not change variable while in iterating
		elif __scope == 'None':
			__scope = [None]
		else:
			__scope = __scope.lower()
			if 'fab' in __scope:
				__scope = ['fabrication']
			elif "install" in __scope or not len(__scope):
				__scope = ['install']
			else:  # executed if there is an invalid value and not blank
				__scope = [None]  # sanitize bad value
			# TODO: silently raise an error

		print __num, __name, __date_due, __date_sent, __gc, __gc_contact, __scope

		try:
			if today() <= __date_due or not __date_sent:
				obj = {'name': __name, 'job_num': __num, 'date_end': __date_due, 'gc': __gc, 'gc_contact': __gc_contact, 'scope': __scope, 'add_to_log': False}
			else:
				obj = {'name': __name, 'job_num': __num, 'date_end': __date_due, 'gc': __gc, 'gc_contact': __gc_contact, 'scope': __scope, 'add_to_log': False, 'completed': __date_sent}
		except TypeError:  # __date_due is 'ASAP'
			if not __date_sent:
				obj = {'name': __name, 'job_num': __num, 'date_end': __date_due, 'gc': __gc, 'gc_contact': __gc_contact, 'scope': __scope, 'add_to_log': False}
			else:
				obj = {'name': __name, 'job_num': __num, 'date_end': __date_due, 'gc': __gc, 'gc_contact': __gc_contact, 'scope': __scope, 'add_to_log': False, 'completed': __date_sent}

		_prev = obj

		yield 'bid', obj
		# TODO: delete leftover variables for debugging purposes


def dump_bids_from_log(estimating_log=environment.get_estimating_log):
	log = load_workbook(estimating_log, read_only=True)
	_sheet = log.get_active_sheet()

	# calculate dimensions to iterate over
	_min_col = 'A'
	_min_row = 3     # do not iterate over header rows
	_max_col = 'I'
	_max_row = _sheet.max_row + 1
	_iter_dim = '%s%s:%s%s' % (_min_col, _min_row, _max_col, _max_row)

	_rows = []
	for _row in _sheet.iter_rows(_iter_dim):
		_rows.append(_row)
	_row_count = len(_rows)
	for _num, _row in zip(range(1, _row_count + 1), _rows):
		_num += 2    # offset for header rows
		yield (_num), _row


def find_bid_in_log(obj, sub_hash=None, estimating_log=environment.get_estimating_log):
	"""
	:param obj: bid object to find in log or to reference for sub bid
	:param sub_hash: if specified, returns row number of sub bid hash from given bid object
	:param estimating_log: pathname for estimating log
	:return: row number of given bid or sub_bid object
	"""
	if sub_hash:
		sub_bid = obj.bids[sub_hash]
		_in_sub = False   # boolean for when the parent bid object has been reached while iterating rows
	_bid_dump = dump_bids_from_log()  # creates a generator object
	for _num, _row in _bid_dump:
		if str(_row[0].value) == str(obj.number):
			if sub_hash:
				_in_sub = True
			else:
				return _num      # returns row number where base bid object occurs
			if sub_hash and _in_sub and sub_bid:
				if str(_row[5].value) == str(sub_bid['gc']):
					return _num  # returns row number where sub bid object occurs
	else:
		return False


def insert_bid_row(obj, estimating_log=environment.get_estimating_log):
	"""
	Inserts blank row after the returned value of `find_bid_in_log`
	:param obj: bid object to insert row after
	:param estimating_log: pathname for estimating log
	:return: returns new row number if operation successful. False if there was an error
	"""
	_row = find_bid_in_log(obj, estimating_log=estimating_log)  # desired row to insert after
	with open(estimating_log, 'r+b') as log:
		_wb = load_workbook(log, guess_types=True)
		_old_ws = _wb.get_active_sheet()
		_old_ws.title += 'old-'
		_new_ws = _wb.create_sheet(0, 'Estimating Log')

		_num = 0   # counter when iterating through old worksheet rows
		_return = None  # row number to return
		_new_ws._styles = _old_ws._styles
		for row in _old_ws.rows:
			_num += 1
			for i in row:
				_col = ord(i.column) - 64  # offset from ASCII char value
				if _col <= 9:  # check to see that column has value
					_new_cell = _new_ws.cell(row=_num, column=_col)
					_new_cell.value = i.value
					_new_cell.style = i.style
					if i.font.bold:
						_new_cell.font = bold_font
					else:
						_new_cell.font = base_font
			if _num == _row:
				_new_ws.append(['' for i in range(0, 9)])
				_num += 1
				_return = _num
		for col, dim in _old_ws.column_dimensions.items():
			if dim.width:
				_new_ws.column_dimensions[col].width = dim.width
		_wb.remove_sheet(_old_ws)
		return _wb, _return


def add_bid_to_log(obj, estimating_log=environment.get_estimating_log):
	#TODO: check for rebid
	log = load_workbook(estimating_log)

	_sheet = log.get_active_sheet()
	_nrow = len(_sheet.rows) + 1
	_row = ['number', '_name', 'date_received', 'bid_date', 'date_sent', 'gc', 'gc_contact' , 'method', 'scope']
	_row[4] = None  # skip 'date_sent' and 'method'
	_row[7] = None

	_row = zip(range(1, len(_row) + 1), _row)
	for col, val in _row:
		try:
			if val:
				_val = obj.__getattribute__(val)
				if hasattr(_val, 'strftime'):
					_val = _val.strftime('%m.%d.%Y')
				elif hasattr(_val, 'sort'):
					_val = ', '.join(sorted(_val))
				_sheet.cell(row=_nrow, column=col, value=_val)
		except Exception as e:
			raise Exception("Unexpected value given when writing %s to (%d,%d): %s" % (str(val), _nrow, col, e.args[0]))

	log.save(estimating_log)
	logger.info("Successfully saved %s to Estimating log" % obj)
	print "Successfully saved %s to Estimating log" % obj


def add_sub_bid_to_log(obj, sub_hash, estimating_log=environment.get_estimating_log):
	""" Adds extra row representing a sub bid to an already existing bid
	:param obj: EstimatingJob object to write to log
	:param sub_hash: new sub bid hash to write to log
	:param estimating_log: pathname for estimating log
	:return: None
	"""
	if find_bid_in_log(obj, sub_hash, estimating_log):
		# TODO: ensure that bid attributes shown in log are up to date
		return True
	else:  # bid has not been added to the estimating log yet
		wb, _row_int = insert_bid_row(obj, estimating_log)   # stores new row
		ws = wb.get_active_sheet()

		_attr = (None, None, 'date_received', 'bid_date', None, 'gc', 'gc_contact' , None, 'scope')
		for attr in _attr:
			if attr:  # Do not write to first 2 columns ('number', '_name'), and 'completed' and 'method' columns
				_col = _attr.index(attr) + 1

				_content = obj.bids[sub_hash][attr]

				# format content for Excel sheet
				if attr in ('date_received',
							'bid_date'):  # format datetime content
					try:
						_content = _content.strftime(format='%m.%d.%y')
					except AttributeError:
						pass
					if attr == 'bid_date':
						ws.cell(row=_row_int, column=_col).value = _content
						ws.cell(row=_row_int, column=_col).font = bold_font
						continue  # continue onto next attribute

				elif attr == 'scope':     # separate scope list into string
					_content = ', '.join(_content)

				ws.cell(row=_row_int, column=_col).value = _content
				ws.cell(row=_row_int, column=_col).font = base_font
				# TODO: style row according to bid status

		wb.save(estimating_log)
		return find_bid_in_log(obj, sub_hash, estimating_log)


def update_bid_in_log(obj=None, attr=None, value=None, estimating_log=environment.get_estimating_log, save=True):
	"""
	:param estimatingLog: estimating log file object to write to
	:param obj: object to reflect changes on
	:param attr: object attribute that has been changed
	:param value: object attribute new value
	:return: True if operation successful
	"""
	_attr = ('number', '_name', 'date_received', 'date_end', 'completed', 'gc', 'gc_contact', 'method', 'scope')
	if attr in _attr and hasattr(obj, 'number'):
		if hasattr(estimating_log, 'get_sheet_by_name'):
			log = estimating_log
		else:
			try:
				log = load_workbook(estimating_log, guess_types=True)
			except IOError:
				print "Cannot open PO log with path '%s'" % estimating_log
				return False

		# Assume that the active sheet is the one that we need
		_sheet = log.get_active_sheet()
		_row_count = 0

		for _row in _sheet.rows:
			_row_count += 1
			if _row[0].value == obj.number:
				# column to edit
				_col = _attr.index(attr) + 1

				# update information
				_sheet.cell(row=_row_count, column=_col).value = value

				# TODO: confirm update

				if save:
					log.save(estimating_log)

				return True


def delete_bid_from_log(obj=None, estimating_log=environment.get_estimating_log):
	"""
	:param obj: Object to locate and delete from log
	:param estimating_log: Default file to parse as log
	:return: True if operation successful. False if not
	"""
	return NotImplemented