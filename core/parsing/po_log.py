from datetime import datetime
import hashlib
import os
import shutil
import unicodedata
from parse import parse

import openpyxl
from xlrd import xldate_as_tuple

from core import environment
from core.log import logger


def parse_po_log(po_log=environment.get_po_log):
	# TODO: add logger debugging hooks
	log = openpyxl.load_workbook(po_log, read_only=True)
	_nsheet = len(log.get_sheet_names())
	for _sheetNum in range(1, _nsheet):  ## skip over first worksheet
		_sheet = log.get_sheet_by_name(log.get_sheet_names()[_sheetNum])
		logger.debug('Working on worksheet "%s"' % _sheet.title)

		try:
			_job = [i for i in parse("{} - {}", _sheet.title)]
			yield 'job', _job
		except TypeError:
			print "Encountered error parsing %s" % _sheet.title
			continue  # Assume sheet is not a job

		for _row in _sheet.rows:
			__po = _row[0].value
			logger.debug('Processing row "%s"' % __po)
			try:
				__po = [i for i in parse("{:d}-{}-{:d}", __po)]
			except TypeError:
				logger.debug('...skipped row')
				continue
			if len(__po) is not 3:  # strip Job number, PO prefix, and PO #
				logger.debug('...skipped row')
				# skips empty row by assuming that a row w/o PO is an empty row
				continue

			# TODO: parse vendor cell and create vendor object
			__vend = _row[1].value
			__price = _row[2].value

			if type(_row[3].value) is datetime:  # parse date when PO was issued
				__date_issued = _row[3].value
			else:
				try:
					__date_issued = datetime(*xldate_as_tuple(_row[3].value, 0))
				except ValueError:
					_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
					for _format in _date_formats:
						try:
							__date_issued = datetime.strptime(_row[3].value, _format)
							break
						except ValueError:
							continue
						except TypeError:
							break
			try:  # ensures that __date_issued is in the namespace
				__date_issued
			except NameError:
				__date_issued   = None
			# SKIP _row[4] -> "date expected"

			try:  # parse material list document cell
				__mat_list_val  = unicodedata.normalize('NFKD', _row[5].value).encode('ascii','ignore')
			except TypeError:
				__mat_list_val  = ''
			try:  # parse material list quote cell
				__quote_val     = unicodedata.normalize('NFKD', _row[6].value).encode('ascii','ignore')
			except TypeError:
				__quote_val = ''

			try:  # parse comment cell
				__comment = _row[7].value
			except IndexError:
				__comment = ""


			_return = {}  # Dict to create objects from

			# Job attribute values

			# MaterialList attribute values
			if '\\' in __mat_list_val:
				__mat_list_val = str(__mat_list_val).replace('\\', '/')
				_mat_list = {'doc': os.path.split(__mat_list_val)[1], 'date_sent': __date_issued, 'task': False}  # store material list values

			else:
				_mat_list = {'items': __mat_list_val, 'date_sent': __date_issued, 'task': False}
			_return['mat_list'] = _mat_list

			# MaterialListQuote attribute values minus MaterialList object
			if '\\' in __quote_val:
				_quote = {'price': __price, 'vend': __vend, 'date_uploaded': __date_issued, 'doc': os.path.split(__quote_val)[1]}
			else:
				_quote = {'price': __price, 'vend': __vend, 'date_uploaded': __date_issued}
			_return['list_quote'] = _quote

			# PO log attributes
			_po_num  = __po[2]
			_po = {'date_issued': __date_issued, 'desc':__comment, 'po_num': _po_num, 'update': False}
			_return['po'] = _po

			yield 'po', _return


def find_job_in_log(obj, po_log=environment.get_po_log):
	"""
	Finds and returns the sheet title for the passed job object in the given po log
	:param obj: jobs object to find in log
	:param po_log: path to PO log to parse
	:return: integer which has the po log for given jobs
	"""
	if hasattr(po_log, 'get_sheet_by_name'):
		# checks to see if Book object was passed
		log = po_log
	else:
		log = openpyxl.load_workbook(po_log, read_only=True)

	if obj is not str:
		_sheet_name = '%d - %s' % (obj.number, obj._name)
	else:
		_sheet_name = obj
	try:
		_sheet = log.get_sheet_by_name(_sheet_name)
		return _sheet, log
	except KeyError:
		return False


def add_job_to_log(obj, po_log=environment.get_po_log, save=True):
	"""
	Adds a spreadsheet page to the passed log file to log POs for the given jobs
	:param obj: job object or int to add to log. Function assumes that the AwardedJob is new to the company
	:param po_log: path to PO log to to add spreadsheet page to
	:return: returns output of find_job_in_log to confirm output
	"""
	if hasattr(po_log, 'get_sheet_by_name'):
		# checks to see if Book object was passed
		log = po_log
	else:
		_po_log = os.path.split(po_log)
		_po_log = os.path.join(_po_log[0], '_%s' % _po_log[1])
		shutil.copy2(po_log, _po_log)
		log = openpyxl.load_workbook(_po_log)

	_sheet_name = obj.sheet_name

	try:
		# returns False if sheet already exists
		log.get_sheet_by_name(_sheet_name)
		return False
	except KeyError:
		_sheet = log.create_sheet(index=int(-2), title=_sheet_name)
		# TODO: add header rows to _sheet
		if save:
			log.save(po_log)
		return find_job_in_log(obj, log)


def dump_pos_from_log(job, po_log=environment.get_po_log):

	_sheet, log = find_job_in_log(job, po_log=po_log)

	# calculate dimensions to iterate over
	_min_col = 'A'
	_min_row = 3     # do not iterate over header rows
	_max_col = 'I'
	_max_row = _sheet.max_row + 1
	_iter_dim = '%s%s:%s%s' % (_min_col, _min_row, _max_col, _max_row)

	_rows = []
	for _row in _sheet.iter_rows(_iter_dim):
		_rows.append(_row)
	_row_count = len(_rows)
	for _num, _row in zip(range(1, _row_count + 1), _rows):
		_num += 2    # offset for header rows
		yield (_num), _row


def find_po_in_log(obj, po_log=environment.get_po_log):
	"""
	Finds and returns spreadsheet page, and row number which corresponds to the given PO object passed
	:param obj: PO object to find in spreadsheet
	:param po_log: path to PO log to parse
	:return: returns tuple containing spreadsheet page integer and row number integer
	"""
	if hasattr(obj, 'job'):
		_job = obj.job

		_sheet, po_log = find_job_in_log(_job, po_log)
		_po_dump = dump_pos_from_log(_job, po_log)  # creates a generator object
		for _num, _row in _po_dump:
			if str(_row[0].value) == str(obj.name):
				_po_row = _num
				return _sheet.title, _po_row


def add_po_to_log(obj, po_log=environment.get_po_log, save=True):
	if hasattr(po_log, 'get_sheet_by_name'):
		log = po_log
	else:
		try:
			log = openpyxl.load_workbook(po_log, guess_types=True)
		except IOError:
			print "'%s' is not a valid file path. Cannot update PO log." % po_log
			return False

	_sheet_name = obj.job.sheet_name
	_sheet = log.get_sheet_by_name(_sheet_name)
	# TODO: implement algorithm to apply styling and organization to PO log
	_nrow = len(_sheet.rows) + 1
	if _nrow < 3:
		_nrow = 3

	# iterate through all rows and cache all POs on worksheet to validate obj against
	_pos = []
	for _row in _sheet.rows:
		_po = _row[0].value
		_pos.append(_po)

	if obj.__repr__() not in _pos:
		_date_issued = obj.date_issued.strftime("%m.%d.%y")
		try:
			_mdoc = os.path.join(*obj.mat_list.doc)
		except TypeError:
			_mdoc = ''
		try:
			_qdoc = os.path.join(*obj.quote.doc)
		except TypeError:
			_qdoc = ''
		_row = (obj.name, obj.vend, obj.price, _date_issued, None, _mdoc, _qdoc, None, None)
		_row = zip(range(1, len(_row) + 1), _row)
		for col, val in _row:
			try:
				_sheet.cell(row=_nrow, column=col, value=str(val))
			except Exception as e:
				raise Exception("Unexpected value given when writing %s to (%d,%d): %s" % (str(val), _nrow, col, e.args[0]))
		if save:
			log.save(po_log)
			logger.info("Successfully saved %s to PO log" % obj)
			print "Successfully saved %s to PO log" % obj
	return find_po_in_log(obj, po_log)


def update_po_in_log(obj=None, attr=None, value=None, po_log=environment.get_po_log, save=True):
	"""
	:param po_log: po_log file object to write to
	:param obj: object to reflect changes on
	:param attr: object attribute that has been changed
	:param value: object attribute new value
	:return: True if operation successful
	"""
	_attr = ('number', 'vend', 'price', 'date_sent', 'date_expected', 'mat_list', 'quote', 'ordered_by', 'quote_id')
	if attr in _attr:
		if hasattr(po_log, 'get_sheet_by_name'):
			log = po_log
		else:
			try:
				_po_log = os.path.split(po_log)
				_po_log = os.path.join(_po_log[0], '_%s' % _po_log[1])
				os.rename(po_log, _po_log)
				log = openpyxl.load_workbook(_po_log, guess_types=True)
			except IOError:
				print "'%s' is not a valid file path. Cannot update PO log." % po_log
				return False

		# pass opened PO Log object to function
		_sheet_name, _row = find_po_in_log(obj, log)
		_sheet = log.get_sheet_by_name(_sheet_name)

		# column to edit
		_col = _attr.index(attr) + 1

		# update information
		_sheet.cell(row=_row, column=_col).value = value

		# TODO: confirm update

		if save:
			log.save(po_log)

		return True


def get_po_attr(obj, attr, po_log=environment.get_po_log):
	# TODO: optimize this bs initialization
	_attr = {'number': 'A', 'vend': 'B', 'price': 'C', 'date_sent': 'D', 'date_expected': 'E',
	         'mat_list': 'F', 'quote': 'J', 'ordered_by': 'K', 'quote_id': 'L'}
	if attr in _attr:
		if hasattr(po_log, 'get_sheet_by_name'):
			log = po_log
		else:
			log = openpyxl.load_workbook(po_log, guess_types=True)

		_sheet, _row = find_po_in_log(obj, po_log)
		_sheet = log.get_sheet_by_name(_sheet)

		cell = '%s%d' % (_attr[attr], _row)
		val = _sheet[cell].value

		return val


def set_log_style(po_log=environment.get_po_log):
	log = openpyxl.load_workbook(po_log)

	_cols = (('number', 'A', 18),    # automatic width
	         ('vend', 'B', 17.5),
	         ('price', 'C', 10),
	         ('date_sent', 'D', 12),
	         ('date_expected', 'E', 11),
	         ('mat_list', 'F', 45),
	         ('quote', 'G', 60),
	         ('ordered_by', 'H', 15),
	         ('comments', 'I', 10))

	_sheets = log.get_sheet_names()
	for _sheet in _sheets:
		ws = log.get_sheet_by_name(_sheet)
		for col in _cols:
			try:
				ws.column_dimensions[col[1]].width = col[2]
			except KeyError:
				continue

	log.save(po_log)


def check_po_log(po_log=environment.get_po_log):
	""" Checks to see that content in database is the updated based on the stored hash of po_log file """
	md5 = hashlib.md5()
	with open(po_log, 'rb') as _po_log:
		buf = _po_log.read()
		md5.update(buf)
		_hash = str(md5.hexdigest())
		if _hash != environment.last_po_log_hash:
			logger.info('updating PO Log hash digest stored')
			environment.set_po_log_hash(_hash)
			return False
		return True