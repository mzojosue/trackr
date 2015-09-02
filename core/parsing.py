import hashlib
from parse import parse
from xlrd import xldate_as_tuple
import openpyxl

from datetime import datetime, timedelta
import unicodedata
import os, shutil

import objects
import environment
from log import logger
from core.scheduler import scheduler


def check_po_log(po_log=environment.get_po_log):
	""" Checks to see that content in database is the updated based on the stored hash of po_log file """
	md5 = hashlib.md5()
	with open(po_log, 'rb') as _po_log:
		buf = _po_log.read()
		md5.update(buf)
		_hash = str(md5.hexdigest())
		if _hash != environment.last_po_log_hash:
			logger.info('updating PO Log hash digest stored')
			environment.set_po_log_hash(_hash)
			return False
		return True

def import_job_info(jobInfo):
	return NotImplemented



# Estimating Log Methods #

def check_estimating_log(estimating_log=environment.get_estimating_log):
	""" Checks to see that content in database is the updated based on the stored hash of po_log file """
	md5 = hashlib.md5()
	with open(estimating_log, 'rb') as _estimating_log:
		buf = _estimating_log.read()
		md5.update(buf)
		_hash = str(md5.hexdigest())
		if _hash != environment.last_estimating_log_hash:
			logger.info('updating Estimating Log hash digest stored')
			environment.set_estimating_log_hash(_hash)
			return False
		return True

def import_estimating_log(estimating_log=environment.get_estimating_log):
	log = openpyxl.load_workbook(estimating_log, read_only=True)
	_sheet = log.get_active_sheet()
	logger.debug('Opening Estimating Log')

	_prev = None  # buffer for storing previous bids. Used for grouping and alternate bidders
	for _row in _sheet.rows:
		__num = _row[0].value
		print "Processing bid row %s" % __num

		try:
			__num = int(__num)
			__name = unicodedata.normalize('NFKD', _row[1].value).encode('ASCII', 'ignore')
		except (ValueError, TypeError):
			# attempt to parse _row as a sub bid
			if _row[5].value and __num == None:  # sub bid hash is based off of 'gc' string/column and 'number' should be None
				_attr = (None, None, 'date_received', 'bid_date', None, 'gc', 'gc_contact' , None, 'scope')
				sub_bid = {}  # stores grabbed values
				for attr in _attr:
					if attr:
						_col = _attr.index(attr)
						sub_bid[attr] = _row[_col].value

				for i in ('date_received', 'bid_date'):
					__date = sub_bid[i]
					if "@" in str(__date):
						__date_due = [i for i in parse("{} @ {}", __date)]
						# create datetime object
					else:
						__date = [__date]
					_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
					for _format in _date_formats:
						try:
							__due = datetime.strptime(__date[0], _format)
							if __date[1:]:
								#TODO: implement
								#TODO: translate AM/PM
								#__due.hour = parse("{:d}{}", __date_due[1])[0]
								pass
							sub_bid[i] = __due
							break
						except ValueError:
							continue
						except TypeError:
							break
				if hasattr(_prev, 'bids'):
					_prev.add_sub(add_to_log=False, **sub_bid)
			continue  # move onto next row

		__date_recvd = _row[2].value
		if __date_recvd is not None:
			_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
			for _format in _date_formats:
				try:
					__date_recvd = datetime.strptime(__date_recvd, _format)
					break
				except ValueError:
					continue
				except TypeError:
					break

		# Parse Bid Due Date #
		__date_due = _row[3].value
		if __date_due is None:
			__date_due = 'ASAP'
		elif __date_due == 'ASAP':
			__date_due = str('ASAP')
		elif __date_due != 'ASAP':
			try:
				__date_due = _row[3].value
				__date_due.date()
			except AttributeError:
				if "@" in str(__date_due):
					__date_due = [i for i in parse("{} @ {}", __date_due)]
				# create datetime object
				else:
					__date_due = [__date_due]
				_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
				for _format in _date_formats:
					try:
						__due = datetime.strptime(__date_due[0], _format)
						if __date_due[1:]:
							#TODO: implement
							#TODO: translate AM/PM
							#__due.hour = parse("{:d}{}", __date_due[1])[0]
							pass
						__date_due = __due
						break
					except ValueError:
						continue
					except TypeError:
						break
		# Parse Date Submitted #
		__date_sent = _row[4].value
		if hasattr(__date_sent, 'lower') and __date_sent.lower() == "no bid":
			__date_sent = "No bid"
		elif __date_sent is not None:
			_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
			for _format in _date_formats:
				try:
					__date_sent = datetime.strptime(__date_sent, _format)
					break
				except ValueError:
					continue
				except TypeError:
					break

		if _row[5].value: __gc = _row[5].value          # Default: None
		else: __gc = None
		if _row[6].value: __gc_contact = _row[6].value  # Default: None
		else: __gc_contact = None
		#  __via = _row[7].value# default: email

		# Scope parsing #
		__scope = str(_row[8].value)
		if ',' in __scope or len(__scope) == 1:
			_tmp_scope = []
			valid_scope = ('M', 'E', 'I', 'B', 'P')
			for _letter in __scope:
				if _letter in valid_scope:
					_tmp_scope.append(_letter)
			__scope = _tmp_scope   # does not change variable while in iterating
		elif __scope == 'None':
			__scope = [None]
		else:
			__scope = __scope.lower()
			if 'fab' in __scope:
				__scope = ['fabrication']
			elif "install" in __scope or not len(__scope):
				__scope = ['install']
			else:  # executed if there is an invalid value and not blank
				__scope = [None]  # sanitize bad value
				# TODO: silently raise an error

		print __num, __name, __date_due, __date_sent, __gc, __gc_contact, __scope
		try:
			if objects.today() <= __date_due or not __date_sent:
				obj = objects.EstimatingJob(__name, __num, date_end=__date_due, gc=__gc, gc_contact=__gc_contact, scope=__scope, add_to_log=False)
			else:
				obj = objects.EstimatingJob(__name, __num, date_end=__date_due, gc=__gc, gc_contact=__gc_contact, scope=__scope, completed=__date_sent, add_to_log=False)
		except TypeError:  # __date_due is 'ASAP'
			if not __date_sent:
				obj = objects.EstimatingJob(__name, __num, date_end=__date_due, gc=__gc, gc_contact=__gc_contact, scope=__scope, add_to_log=False)
			else:
				obj = objects.EstimatingJob(__name, __num, date_end=__date_due, gc=__gc, gc_contact=__gc_contact, scope=__scope, completed=__date_sent, add_to_log=False)
		_prev = obj
		# TODO: delete leftover variables for debugging purposes

def dump_bids_from_log(estimating_log=environment.get_estimating_log):
	log = openpyxl.load_workbook(estimating_log, read_only=True)
	_sheet = log.get_active_sheet()

	# calculate dimensions to iterate over
	_min_col = 'A'
	_min_row = 3     # do not iterate over header rows
	_max_col = 'I'
	_max_row = _sheet.max_row + 1
	_iter_dim = '%s%s:%s%s' % (_min_col, _min_row, _max_col, _max_row)

	_rows = []
	for _row in _sheet.iter_rows(_iter_dim):
		_rows.append(_row)
	_row_count = len(_rows)
	for _num, _row in zip(range(1, _row_count + 1), _rows):
		_num += 2    # offset for header rows
		yield (_num), _row

def find_bid_in_log(obj, sub_hash=None, estimating_log=environment.get_estimating_log):
	"""
	:param obj: bid object to find in log or to reference for sub bid
	:param sub_hash: if specified, returns row number of sub bid hash from given bid object
	:param estimating_log: pathname for estimating log
	:return: row number of given bid or sub_bid object
	"""
	if sub_hash:
		sub_bid = obj.bids[sub_hash]
		_in_sub = False   # boolean for when the parent bid object has been reached while iterating rows
	_bid_dump = dump_bids_from_log()  # creates a generator object
	for _num, _row in _bid_dump:
		if str(_row[0].value) == str(obj.number):
			if sub_hash:
				_in_sub = True
			else:
				return _num      # returns row number where base bid object occurs
			if sub_hash and _in_sub and sub_bid:
				if str(_row[5].value) == str(sub_bid['gc']):
					return _num  # returns row number where sub bid object occurs
	else:
		return False

def insert_bid_row(obj, estimating_log=environment.get_estimating_log):
	"""
	Inserts blank row after the returned value of `find_bid_in_log`
	:param obj: bid object to insert row after
	:param estimating_log: pathname for estimating log
	:return: returns new row number if operation successful. False if there was an error
	"""
	_row = find_bid_in_log(obj, estimating_log=estimating_log)  # desired row to insert after
	with open(estimating_log, 'r+b') as log:
		_wb = openpyxl.load_workbook(log, guess_types=True)
		_old_ws = _wb.get_active_sheet()
		_old_ws.title = 'old-' + _old_ws.title
		_new_ws = _wb.create_sheet(0, 'Estimating Log')

		_num = 0   # counter when iterating through old worksheet rows
		_return = None  # row number to return
		for row in _old_ws.rows:
			_num += 1
			for i in row:
				_col = ord(i.column) - 64  # offset from ASCII char value
				_new_ws.cell(row=_num, column=_col, value=i.value)
				_new_ws.cell(row=_num, column=_col).style = i.style
			if _num == _row:
				_new_ws.append([''])
				_num += 1
				_return = _num
		_new_ws._styles = _old_ws._styles
		for col, dim in _old_ws.column_dimensions.items():
			if dim.width:
				_new_ws.column_dimensions[col].width = dim.width
		_wb.remove_sheet(_old_ws)
		return (_wb, _return)

def add_bid_to_log(obj, estimating_log=environment.get_estimating_log):
	#TODO: check for rebid
	if type(obj) is int:
		obj = objects.EstimatingJob.find(obj)
	log = openpyxl.load_workbook(estimating_log)

	_sheet = log.get_active_sheet()
	_nrow = len(_sheet.rows) + 1
	_row = ['number', '_name', 'date_received', 'bid_date', 'date_sent', 'gc', 'gc_contact' , 'method', 'scope']
	_row[4] = None  # skip 'date_sent' and 'method'
	_row[7] = None

	_row = zip(range(1, len(_row) + 1), _row)
	for col, val in _row:
		try:
			if val:
				_val = obj.__getattribute__(val)
				if hasattr(_val, 'strftime'):
					_val = _val.strftime('%m.%d.%Y')
				elif hasattr(_val, 'sort'):
					_val = ', '.join(sorted(_val))
				_sheet.cell(row=_nrow, column=col, value=_val)
		except Exception as e:
			raise Exception("Unexpected value given when writing %s to (%d,%d): %s" % (str(val), _nrow, col, e.args[0]))

	log.save(estimating_log)
	logger.info("Successfully saved %s to Estimating log" % obj)
	print "Successfully saved %s to Estimating log" % obj

def add_sub_bid_to_log(obj, sub_hash, estimating_log=environment.get_estimating_log):
	""" Adds extra row representing a sub bid to an already existing bid
	:param obj: EstimatingJob object to write to log
	:param sub_hash: new sub bid hash to write to log
	:param estimating_log: pathname for estimating log
	:return: None
	"""
	if find_bid_in_log(obj, sub_hash, estimating_log):
		# TODO: ensure that bid attributes shown in log are up to date
		return True
	else:  # bid has not been added to the estimating log yet
		wb, _row_int = insert_bid_row(obj, estimating_log)   # stores new row
		ws = wb.get_active_sheet()

		_attr = (None, None, 'date_received', 'bid_date', None, 'gc', 'gc_contact' , None, 'scope')
		for attr in _attr:
			if attr:  # Do not write to first 2 columns ('number', '_name') and 'method' column
				_col = _attr.index(attr) + 1

				_content = obj.bids[sub_hash][attr]

				# format content for Excel sheet
				if attr in ('date_received',
							'bid_date'):  # format datetime content
					_content = _content.strftime(format='%m.%d.%y')
				elif attr is 'scope':     # separate scope list into string
					_content = ', '.join(_content)

				ws.cell(row=_row_int, column=_col).value = _content
				# TODO: style element according to bid status

		# TODO: confirm update

		wb.save(estimating_log)
		return True

def update_bid_in_log(obj=None, attr=None, value=None, estimating_log=environment.get_estimating_log, save=True):
	"""
	:param estimatingLog: estimating log file object to write to
	:param obj: object to reflect changes on
	:param attr: object attribute that has been changed
	:param value: object attribute new value
	:return: True if operation successful
	"""
	_attr = ('number', '_name', 'date_received', 'date_end', 'completed', 'gc', 'gc_contact', 'method', 'scope')
	if attr in _attr and hasattr(obj, 'number'):
		if hasattr(estimating_log, 'get_sheet_by_name'):
			log = estimating_log
		else:
			try:
				_est_log = os.path.split(estimating_log)
				_est_log = os.path.join(_est_log[0], '_%s' % _est_log[1])
				#os.rename(estimating_log, _est_log)
				#log = openpyxl.load_workbook(_est_log, guess_types=True)
				log = openpyxl.load_workbook(estimating_log, guess_types=True)
			except IOError:
				print "'%s' is not a valid file path. Cannot update PO log." % estimating_log
				return False

		# Assume that the active sheet is the one that we need
		_sheet = log.get_active_sheet()
		_row_count = 0

		for _row in _sheet.rows:
			_row_count += 1
			if _row[0].value == obj.number:
				# column to edit
				_col = _attr.index(attr) + 1

				# update information
				_sheet.cell(row=_row_count, column=_col).value = value

				# TODO: confirm update

				if save:
					log.save(estimating_log)

				return True

def delete_bid_from_log(obj=None, estimating_log=environment.get_estimating_log):
	"""
	:param obj: Object to locate and delete from log
	:param estimating_log: Default file to parse as log
	:return: True if operation successful. False if not
	"""
	return NotImplemented



# PO Log functions #

def import_po_log(create=False, po_log=environment.get_po_log):
	# TODO: add logger debugging hooks
	log = openpyxl.load_workbook(po_log, read_only=True)
	_nsheet = len(log.get_sheet_names()) - 2    # skip over "Blank Page" and "Small Projects" pages
	for _sheetNum in range(1, _nsheet):
		_sheet = log.get_sheet_by_name(log.get_sheet_names()[_sheetNum])
		logger.debug('Working on worksheet "%s"' % _sheet.title)
		if create:
			_job = objects.AwardedJob(*[i for i in parse("{}-{}", _sheet.title)])
		for _row in _sheet.rows:
			__po = _row[0].value
			logger.debug('Processing row "%s"' % __po)
			try:
				__po = [i for i in parse("{:d}-{}-{:d}", __po)]
			except TypeError:
				logger.debug('...skipped row')
				continue
			if len(__po) is not 3:
				logger.debug('...skipped row')
				# skips empty row by assuming that a row
				continue
			# TODO: parse vendor cell and create vendor object
			__vend = _row[1].value
			__price = _row[2].value
			if type(_row[3].value) is datetime:
				__date_issued = _row[3].value
			else:
				try:
					__date_issued = datetime(*xldate_as_tuple(_row[3].value, 0))
				except ValueError:
					_date_formats = ['%m.%d.%y', '%m.%d.%Y', '%m/%d/%y', '%m/%d/%Y']
					for _format in _date_formats:
						try:
							__date_issued = datetime.strptime(_row[3].value, _format)
							break
						except ValueError:
							continue
						except TypeError:
							break
			try:
				__date_issued
			except NameError:
				__date_issued   = None

			# SKIP _row[4] -> "date expected"
			try:
				__mat_list_val  = unicodedata.normalize('NFKD', _row[5].value).encode('ascii','ignore')
			except TypeError:
				__mat_list_val  = ''
			try:
				__quote_val     = unicodedata.normalize('NFKD', _row[6].value).encode('ascii','ignore')
			except TypeError:
				__quote_val = ''

			try:
				__comment = _row[7].value
			except IndexError:
				__comment = ""

			if create:
				# Create MaterialList objects
				if '\\' in __mat_list_val:
					__mat_list_val = str(__mat_list_val).replace('\\', '/')
					_mat_list = objects.MaterialList(job=_job, doc=objects.os.path.split(__mat_list_val), date_sent=__date_issued, task=False)

				else:
					_mat_list = objects.MaterialList(job=_job, items=__mat_list_val, date_sent=__date_issued, task=False)
				_mat_list.sent_out = True
				if _mat_list.age > 5:
					_mat_list.delivered = True

				# Create MaterialListQuote objects
				if '\\' in __quote_val:
					_quote = objects.MaterialListQuote(mat_list=_mat_list, price=__price, vend=__vend, date_uploaded=__date_issued, doc=objects.os.path.split(__quote_val))
				else:
					_quote = objects.MaterialListQuote(mat_list=_mat_list, price=__price, vend=__vend, date_uploaded=__date_issued)

				# Create PO objects
				_po_pre  = '-'.join([str(i) for i in __po[:2]])
				_po_num  = __po[2]
				if str(_po_pre) is not str(_job.po_pre):
					_pre = _po_pre
				else:
					_pre = None
				_po = objects.PO(_job, _mat_list, __date_issued, _quote, desc=__comment, po_num=_po_num, po_pre=_pre, update=False)
			del __po, __vend, __price, __date_issued, __mat_list_val, __quote_val, __comment

def find_job_in_log(obj, po_log=environment.get_po_log):
	"""
	Finds and returns the sheet title for the passed job object in the given po log
	:param obj: jobs object to find in log
	:param po_log: path to PO log to parse
	:return: integer which has the po log for given jobs
	"""
	if hasattr(po_log, 'get_sheet_by_name'):
		# checks to see if Book object was passed
		log = po_log
	else:
		log = openpyxl.load_workbook(po_log, read_only=True)

	if type(obj) is int:
		obj = objects.AwardedJob.find(obj)

	_sheet_name = '%d - %s' % (obj.number, obj._name)
	try:
		_sheet = log.get_sheet_by_name(_sheet_name)
		return _sheet, log
	except KeyError:
		return False

def add_job_to_log(obj, po_log=environment.get_po_log, save=True):
	"""
	Adds a spreadsheet page to the passed log file to log POs for the given jobs
	:param obj: job object or int to add to log. Function assumes that the AwardedJob is new to the company
	:param po_log: path to PO log to to add spreadsheet page to
	:return: returns output of find_job_in_log to confirm output
	"""
	if hasattr(po_log, 'get_sheet_by_name'):
		# checks to see if Book object was passed
		log = po_log
	else:
		_po_log = os.path.split(po_log)
		_po_log = os.path.join(_po_log[0], '_%s' % _po_log[1])
		shutil.copy2(po_log, _po_log)
		log = openpyxl.load_workbook(_po_log)
	if type(obj) is int:
		obj = objects.AwardedJob.find(obj)
	_sheet_name = obj.sheet_name

	try:
		# returns False if sheet already exists
		log.get_sheet_by_name(_sheet_name)
		return False
	except KeyError:
		_sheet = log.create_sheet(index=int(-2), title=_sheet_name)
		# TODO: add header rows to _sheet
		if save:
			log.save(po_log)
		return find_job_in_log(obj, log)

def dump_pos_from_log(job, po_log=environment.get_po_log):
	if type(job) is int:
		# TODO: implement dumping POs from CompletedJob object instead of just AwardedJob
		# convert string object to job object
		job = objects.AwardedJob.find(job)

	_sheet, log = find_job_in_log(job, po_log=po_log)

	# calculate dimensions to iterate over
	_min_col = 'A'
	_min_row = 3     # do not iterate over header rows
	_max_col = 'I'
	_max_row = _sheet.max_row + 1
	_iter_dim = '%s%s:%s%s' % (_min_col, _min_row, _max_col, _max_row)

	_rows = []
	for _row in _sheet.iter_rows(_iter_dim):
		_rows.append(_row)
	_row_count = len(_rows)
	for _num, _row in zip(range(1, _row_count + 1), _rows):
		_num += 2    # offset for header rows
		yield (_num), _row

def find_po_in_log(obj, po_log=environment.get_po_log):
	"""
	Finds and returns spreadsheet page, and row number which corresponds to the given PO object passed
	:param obj: PO object to find in spreadsheet
	:param po_log: path to PO log to parse
	:return: returns tuple containing spreadsheet page integer and row number integer
	"""
	if hasattr(obj, 'job'):
		_job = obj.job

		_sheet, po_log = find_job_in_log(_job, po_log)
		_po_dump = dump_pos_from_log(_job, po_log)  # creates a generator object
		for _num, _row in _po_dump:
			if str(_row[0].value) == str(obj.name):
				_po_row = _num
				return _sheet.title, _po_row

def add_po_to_log(obj, po_log=environment.get_po_log, save=True):
	if hasattr(po_log, 'get_sheet_by_name'):
		log = po_log
	else:
		try:
			log = openpyxl.load_workbook(po_log, guess_types=True)
		except IOError:
			print "'%s' is not a valid file path. Cannot update PO log." % po_log
			return False

	_sheet_name = obj.job.sheet_name
	_sheet = log.get_sheet_by_name(_sheet_name)
	# TODO: implement algorithm to apply styling and organization to PO log
	_nrow = len(_sheet.rows) + 1
	if _nrow < 3:
		_nrow = 3

	# iterate through all rows and cache all POs on worksheet to validate obj against
	_pos = []
	for _row in _sheet.rows:
		_po = _row[0].value
		_pos.append(_po)

	if obj.__repr__() not in _pos:
		_date_issued = obj.date_issued.strftime("%m.%d.%y")
		try:
			_mdoc = objects.os.path.join(*obj.mat_list.doc)
		except TypeError:
			_mdoc = ''
		try:
			_qdoc = objects.os.path.join(*obj.quote.doc)
		except TypeError:
			_qdoc = ''
		_row = (obj.name, obj.vend, obj.price, _date_issued, None, _mdoc, _qdoc, None, None)
		_row = zip(range(1, len(_row) + 1), _row)
		for col, val in _row:
			try:
				_sheet.cell(row=_nrow, column=col, value=str(val))
			except Exception as e:
				raise Exception("Unexpected value given when writing %s to (%d,%d): %s" % (str(val), _nrow, col, e.args[0]))
		if save:
			log.save(po_log)
			logger.info("Successfully saved %s to PO log" % obj)
			print "Successfully saved %s to PO log" % obj
	return find_po_in_log(obj, po_log)

def update_po_in_log(obj=None, attr=None, value=None, po_log=environment.get_po_log, save=True):
	"""
	:param po_log: po_log file object to write to
	:param obj: object to reflect changes on
	:param attr: object attribute that has been changed
	:param value: object attribute new value
	:return: True if operation successful
	"""
	_attr = ('number', 'vend', 'price', 'date_sent', 'date_expected', 'mat_list', 'quote', 'ordered_by', 'quote_id')
	if attr in _attr:
		if hasattr(po_log, 'get_sheet_by_name'):
			log = po_log
		else:
			try:
				_po_log = os.path.split(po_log)
				_po_log = os.path.join(_po_log[0], '_%s' % _po_log[1])
				os.rename(po_log, _po_log)
				log = openpyxl.load_workbook(_po_log, guess_types=True)
			except IOError:
				print "'%s' is not a valid file path. Cannot update PO log." % po_log
				return False

		# pass opened PO Log object to function
		_sheet_name, _row = find_po_in_log(obj, log)
		_sheet = log.get_sheet_by_name(_sheet_name)

		# column to edit
		_col = _attr.index(attr) + 1

		# update information
		_sheet.cell(row=_row, column=_col).value = value

		# TODO: confirm update

		if save:
			log.save(po_log)

		return True

def get_po_attr(obj, attr, po_log=environment.get_po_log):
	# TODO: optimize this bs initialization
	_attr = {'number': 'A', 'vend': 'B', 'price': 'C', 'date_sent': 'D', 'date_expected': 'E',
	         'mat_list': 'F', 'quote': 'J', 'ordered_by': 'K', 'quote_id': 'L'}
	if attr in _attr:
		if hasattr(po_log, 'get_sheet_by_name'):
			log = po_log
		else:
			log = openpyxl.load_workbook(po_log, guess_types=True)

		_sheet, _row = find_po_in_log(obj, po_log)
		_sheet = log.get_sheet_by_name(_sheet)

		cell = '%s%d' % (_attr[attr], _row)
		val = _sheet[cell].value

		return val

def set_log_style(po_log=environment.get_po_log):
	log = openpyxl.load_workbook(po_log)

	_cols = (('number', 'A', 18),    # automatic width
	         ('vend', 'B', 17.5),
	         ('price', 'C', 10),
	         ('date_sent', 'D', 12),
	         ('date_expected', 'E', 11),
	         ('mat_list', 'F', 45),
	         ('quote', 'G', 60),
	         ('ordered_by', 'H', 15),
	         ('comments', 'I', 10))

	_sheets = log.get_sheet_names()
	for _sheet in _sheets:
		ws = log.get_sheet_by_name(_sheet)
		for col in _cols:
			try:
				ws.column_dimensions[col[1]].width = col[2]
			except KeyError:
				continue

	log.save(po_log)
